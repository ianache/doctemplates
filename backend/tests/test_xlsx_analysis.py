import io

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from app.services.xlsx_analysis import analyze_xlsx_template


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["A1"] = "Customer: {{cliente.nombre}}"
    worksheet.print_area = "A1:C12"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_analyze_xlsx_template_extracts_sheet_token_and_print_area() -> None:
    analysis = analyze_xlsx_template(_workbook_bytes(), {"cliente.nombre"})

    assert analysis.detected_sheets == [
        {
            "name": "Summary",
            "max_row": 1,
            "max_column": 1,
            "print_area": "'Summary'!$A$1:$C$12",
            "merged_ranges": [],
        }
    ]
    assert analysis.detected_tokens == ["cliente.nombre"]
    assert analysis.validation_warnings == []
    assert analysis.image_slots == []


def test_analyze_xlsx_template_warns_for_unknown_schema_token() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["B4"] = "{{cliente.desconocido}}"
    output = io.BytesIO()
    workbook.save(output)

    analysis = analyze_xlsx_template(output.getvalue(), {"cliente.nombre"})

    assert analysis.validation_warnings[0]["type"] == "unknown_schema_token"
    assert analysis.validation_warnings[0]["cell"] == "B4"
    assert analysis.validation_warnings[0]["sheet"] == "Summary"


def test_analyze_maps_repeat_item_tokens_to_schema_list_fields() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Items"
    worksheet["A2"] = "{{item.name}}"
    worksheet["Z1"] = '[{"sheet":"Items","row":2,"list":"items"}]'
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Items'!$Z$1"))
    output = io.BytesIO()
    workbook.save(output)

    analysis = analyze_xlsx_template(output.getvalue(), {"items[].name"})

    assert analysis.detected_tokens == ["item.name"]
    assert analysis.validation_warnings == []


def test_analyze_warns_for_invalid_repeat_metadata() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Items"
    worksheet["A2"] = "{{item.name}}"
    worksheet["Z1"] = '[{"sheet":"Missing","row":2,"list":"items"}]'
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Items'!$Z$1"))
    output = io.BytesIO()
    workbook.save(output)

    analysis = analyze_xlsx_template(output.getvalue(), {"items[].name"})

    assert analysis.validation_warnings[0]["type"] == "invalid_repeat_metadata"


def test_analyze_detects_image_slots_from_explicit_metadata() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["Z2"] = '[{"sheet":"Summary","cell":"D4","field":"brand.logo"}]'
    workbook.defined_names.add(DefinedName("_docman_images", attr_text="'Summary'!$Z$2"))
    output = io.BytesIO()
    workbook.save(output)

    analysis = analyze_xlsx_template(output.getvalue(), set())

    assert analysis.image_slots == [{"sheet": "Summary", "cell": "D4", "field": "brand.logo"}]
    assert analysis.validation_warnings == []


def test_analyze_warns_for_non_numeric_repeat_row() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Items"
    worksheet["A2"] = "{{item.name}}"
    worksheet["Z1"] = '[{"sheet":"Items","row":"two","list":"items"}]'
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Items'!$Z$1"))
    output = io.BytesIO()
    workbook.save(output)

    analysis = analyze_xlsx_template(output.getvalue(), {"items[].name"})

    assert analysis.validation_warnings[0]["type"] == "invalid_repeat_metadata"
    assert analysis.validation_warnings[0]["message"] == "Repeat metadata row must be numeric."
