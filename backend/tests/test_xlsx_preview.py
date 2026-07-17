import io

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from sqlalchemy.orm import Session as SQLAlchemySession

from app.auth.session_service import create_session
from app.config import settings
from app.models.document_type import DocumentType, DocumentTypeField
from app.models.user import User
from app.models.xlsx_template import XlsxTemplate


def _auth_client(client: TestClient, db_session: SQLAlchemySession) -> User:
    user = User(sub="xlsx-preview-sub", email="xlsx-preview@example.com")
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)
    session = create_session(db_session, user)
    client.cookies.set("docmanagement_session", session.id)
    return user


def _document_type(db_session: SQLAlchemySession, user: User) -> DocumentType:
    document_type = DocumentType(
        name="Preview Workbook",
        description="Preview workbook type",
        allowed_output_formats=["pdf", "xlsx"],
        created_by=user,
        fields=[
            DocumentTypeField(
                name="cliente.nombre",
                type="string",
                description="Customer name",
                position=0,
            )
        ],
    )
    db_session.add(document_type)
    db_session.commit()
    db_session.refresh(document_type)
    return document_type


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["B1"] = "{{cliente.nombre}}"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _repeat_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["A1"] = "{{item.name}}"
    worksheet["Z1"] = '[{"sheet":"Summary","row":1,"list":"items"}]'
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Summary'!$Z$1"))
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _invalid_repeat_metadata_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["A1"] = "Value"
    worksheet["Z1"] = '[{"sheet":"Missing","row":1,"list":"items"}]'
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Summary'!$Z$1"))
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_preview_returns_sheet_cells(
    client: TestClient, db_session: SQLAlchemySession, tmp_path
) -> None:
    user = _auth_client(client, db_session)
    document_type = _document_type(db_session, user)
    original_root = settings.xlsx_template_storage_root
    settings.xlsx_template_storage_root = str(tmp_path)
    try:
        upload_response = client.post(
            "/api/xlsx-templates",
            data={"document_type_id": str(document_type.id), "name": "Preview template"},
            files={
                "file": (
                    "preview.xlsx",
                    _workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert upload_response.status_code == 201
        template_id = upload_response.json()["id"]

        response = client.post(
            f"/api/xlsx-templates/{template_id}/preview",
            json={"mock_data": {"cliente": {"nombre": "ACME"}}},
        )

        assert response.status_code == 200
        body = response.json()
        assert body["sheets"][0]["name"] == "Summary"
        assert body["sheets"][0]["cells"][0]["address"] == "B1"
        assert body["sheets"][0]["cells"][0]["value"] == "ACME"
    finally:
        settings.xlsx_template_storage_root = original_root


def test_preview_uses_template_mock_data_when_body_missing(
    client: TestClient, db_session: SQLAlchemySession, tmp_path
) -> None:
    user = _auth_client(client, db_session)
    document_type = _document_type(db_session, user)
    original_root = settings.xlsx_template_storage_root
    settings.xlsx_template_storage_root = str(tmp_path)
    try:
        template_id = _upload_preview_template(client, document_type.id)
        template = db_session.get(XlsxTemplate, template_id)
        template.mock_data = {"cliente": {"nombre": "Stored"}}
        db_session.commit()

        response = client.post(f"/api/xlsx-templates/{template_id}/preview")

        assert response.status_code == 200
        assert response.json()["sheets"][0]["cells"][0]["value"] == "Stored"
    finally:
        settings.xlsx_template_storage_root = original_root


def test_preview_empty_mock_data_does_not_fall_back_to_template_mock_data(
    client: TestClient, db_session: SQLAlchemySession, tmp_path
) -> None:
    user = _auth_client(client, db_session)
    document_type = _document_type(db_session, user)
    original_root = settings.xlsx_template_storage_root
    settings.xlsx_template_storage_root = str(tmp_path)
    try:
        template_id = _upload_preview_template(client, document_type.id)
        template = db_session.get(XlsxTemplate, template_id)
        template.mock_data = {"cliente": {"nombre": "Stored"}}
        db_session.commit()

        response = client.post(f"/api/xlsx-templates/{template_id}/preview", json={"mock_data": {}})

        assert response.status_code == 200
        assert response.json()["sheets"][0]["cells"][0]["value"] == ""
    finally:
        settings.xlsx_template_storage_root = original_root


def test_preview_returns_400_for_renderer_validation_error(
    client: TestClient, db_session: SQLAlchemySession, tmp_path
) -> None:
    user = _auth_client(client, db_session)
    document_type = _document_type(db_session, user)
    original_root = settings.xlsx_template_storage_root
    settings.xlsx_template_storage_root = str(tmp_path)
    try:
        upload_response = client.post(
            "/api/xlsx-templates",
            data={"document_type_id": str(document_type.id), "name": "Repeat template"},
            files={
                "file": (
                    "repeat.xlsx",
                    _repeat_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert upload_response.status_code == 201

        response = client.post(
            f"/api/xlsx-templates/{upload_response.json()['id']}/preview",
            json={"mock_data": {"items": "not-a-list"}},
        )

        assert response.status_code == 400
        assert response.json()["detail"] == "repeat_list_must_be_array"
    finally:
        settings.xlsx_template_storage_root = original_root


def test_preview_returns_400_for_invalid_repeat_metadata(
    client: TestClient, db_session: SQLAlchemySession, tmp_path
) -> None:
    user = _auth_client(client, db_session)
    document_type = _document_type(db_session, user)
    original_root = settings.xlsx_template_storage_root
    settings.xlsx_template_storage_root = str(tmp_path)
    try:
        upload_response = client.post(
            "/api/xlsx-templates",
            data={"document_type_id": str(document_type.id), "name": "Invalid repeat template"},
            files={
                "file": (
                    "invalid-repeat.xlsx",
                    _invalid_repeat_metadata_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert upload_response.status_code == 201

        response = client.post(f"/api/xlsx-templates/{upload_response.json()['id']}/preview")

        assert response.status_code == 400
        assert response.json()["detail"] == "invalid_repeat_metadata"
    finally:
        settings.xlsx_template_storage_root = original_root


def _upload_preview_template(client: TestClient, document_type_id) -> str:
    response = client.post(
        "/api/xlsx-templates",
        data={"document_type_id": str(document_type_id), "name": "Preview template"},
        files={
            "file": (
                "preview.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201
    return response.json()["id"]
