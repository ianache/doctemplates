from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from app.services.xlsx_images import normalize_image_value
from app.schemas.xlsx_template import XlsxTemplatePreviewResponse
from app.services.xlsx_renderer import preview_xlsx_template, render_xlsx_template


def _summary_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["A1"] = "Customer"
    worksheet["B1"] = "{{cliente.nombre}}"
    worksheet["C1"] = "=SUM(1,2)"
    worksheet.column_dimensions["B"].width = 30
    worksheet.print_area = "A1:C10"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _repeat_workbook_bytes(repeat_json: str = '[{"sheet":"Items","row":2,"list":"items"}]') -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Items"
    worksheet["A1"] = "Name"
    worksheet["B1"] = "Qty"
    worksheet["A2"] = "{{item.name}}"
    worksheet["B2"] = "{{item.qty}}"
    worksheet["Z1"] = repeat_json
    worksheet.column_dimensions["Z"].hidden = True
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Items'!$Z$1"))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_render_replaces_cell_token_and_preserves_workbook_structure() -> None:
    output = render_xlsx_template(_summary_workbook_bytes(), {"cliente": {"nombre": "ACME"}})
    workbook = load_workbook(BytesIO(output), data_only=False)
    worksheet = workbook["Summary"]

    assert worksheet["B1"].value == "ACME"
    assert str(worksheet.print_area) == "'Summary'!$A$1:$C$10"
    assert worksheet.column_dimensions["B"].width == 30
    assert worksheet["C1"].value == "=SUM(1,2)"


def test_preview_serializes_datetime_cells_as_strings() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["A1"] = datetime(2026, 7, 16, 17, 26, 19)
    buffer = BytesIO()
    workbook.save(buffer)

    preview = preview_xlsx_template(buffer.getvalue(), {})

    assert preview["sheets"][0]["cells"][0]["value"] == "2026-07-16 17:26:19"
    XlsxTemplatePreviewResponse(**preview)


def test_render_missing_nested_payload_value_as_empty_string() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "{{contrato.codigo_cliente}}"
    buffer = BytesIO()
    workbook.save(buffer)

    output = render_xlsx_template(buffer.getvalue(), {})

    worksheet = load_workbook(BytesIO(output)).active
    assert worksheet["A1"].value is None


def test_render_escapes_payload_formula_values() -> None:
    output = render_xlsx_template(_summary_workbook_bytes(), {"cliente": {"nombre": "=HYPERLINK(\"x\")"}})
    workbook = load_workbook(BytesIO(output), data_only=False)

    assert workbook["Summary"]["B1"].value == "'=HYPERLINK(\"x\")"


def test_render_repeats_explicit_row_for_list_items() -> None:
    output = render_xlsx_template(
        _repeat_workbook_bytes(),
        {"items": [{"name": "A", "qty": 1}, {"name": "B", "qty": 2}]},
    )
    rendered = load_workbook(BytesIO(output))["Items"]

    assert rendered["A2"].value == "A"
    assert rendered["B2"].value == "1"
    assert rendered["A3"].value == "B"
    assert rendered["B3"].value == "2"


def test_empty_repeat_row_updates_later_repeat_offsets() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Items"
    worksheet["A2"] = "{{item.name}}"
    worksheet["A3"] = "{{item.name}}"
    worksheet["Z1"] = (
        '[{"sheet":"Items","row":2,"list":"empty_items"},'
        '{"sheet":"Items","row":3,"list":"items"}]'
    )
    worksheet.column_dimensions["Z"].hidden = True
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Items'!$Z$1"))
    buffer = BytesIO()
    workbook.save(buffer)

    output = render_xlsx_template(buffer.getvalue(), {"empty_items": [], "items": [{"name": "A"}]})
    rendered = load_workbook(BytesIO(output))["Items"]

    assert rendered["A2"].value == "A"
    assert rendered["A3"].value is None


def test_render_rejects_merged_cells_intersecting_repeat_row() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Items"
    worksheet["A2"] = "{{item.name}}"
    worksheet.merge_cells("A2:B2")
    worksheet["Z1"] = '[{"sheet":"Items","row":2,"list":"items"}]'
    workbook.defined_names.add(DefinedName("_docman_repeats", attr_text="'Items'!$Z$1"))
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(ValueError, match="unsupported_merge_in_repeat_range"):
        render_xlsx_template(buffer.getvalue(), {"items": [{"name": "A"}]})


def test_normalize_png_data_url() -> None:
    data_url = (
        "data:image/png;base64,"
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )

    image = normalize_image_value(data_url)

    assert image.mime_type == "image/png"
    assert image.width >= 1
    assert image.height >= 1


def test_render_inserts_explicitly_anchored_image() -> None:
    data_url = (
        "data:image/png;base64,"
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )

    output = render_xlsx_template(_summary_workbook_bytes(), {}, {"Summary!D4": data_url})
    workbook = load_workbook(BytesIO(output))

    assert len(workbook["Summary"]._images) == 1


def test_normalize_rejects_invalid_base64_data_url() -> None:
    with pytest.raises(ValueError, match="invalid_image_payload"):
        normalize_image_value("data:image/png;base64,not-valid-base64")
