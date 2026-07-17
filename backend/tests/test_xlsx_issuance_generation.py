from io import BytesIO
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.services.document_generation import XLSX_MIME_TYPE, generate_document_file


class _Storage:
    def __init__(self, workbook_bytes: bytes) -> None:
        self.workbook_bytes = workbook_bytes

    def get(self, key: str, category: str) -> bytes:
        assert key == "template.xlsx"
        assert category == "xlsx_templates"
        return self.workbook_bytes


class _Template:
    storage_key = "template.xlsx"
    image_slots = []


class _Design:
    output_format = "xlsx"
    xlsx_template = _Template()


class _Issuance:
    id = uuid4()
    design_version = _Design()
    input_data = {"cliente": {"nombre": "ACME"}}


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.active["A1"] = "{{cliente.nombre}}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_xlsx_mime_type_constant() -> None:
    assert XLSX_MIME_TYPE == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_generate_document_file_returns_pdf_document(monkeypatch) -> None:
    class PdfDesign:
        output_format = "pdf"

    class PdfIssuance:
        id = uuid4()
        design_version = PdfDesign()
        input_data = {"name": "ACME"}

    monkeypatch.setattr(
        "app.services.document_generation.generate_composed_pdf",
        lambda *args, **kwargs: b"%PDF-1.4",
    )

    issuance = PdfIssuance()
    generated = generate_document_file(issuance, db=None, storage_provider=_Storage(_workbook_bytes()))

    assert generated.content == b"%PDF-1.4"
    assert generated.mime_type == "application/pdf"
    assert generated.filename == f"{issuance.id}.pdf"


def test_generate_document_file_renders_xlsx_workbook() -> None:
    issuance = _Issuance()
    generated = generate_document_file(issuance, db=None, storage_provider=_Storage(_workbook_bytes()))

    assert generated.mime_type == XLSX_MIME_TYPE
    assert generated.filename == f"{issuance.id}.xlsx"
    assert generated.extension == "xlsx"
    workbook = load_workbook(BytesIO(generated.content))
    assert workbook.active["A1"].value == "ACME"


def test_generate_document_file_inserts_xlsx_images_from_slots() -> None:
    class TemplateWithImage:
        storage_key = "template.xlsx"
        image_slots = [{"sheet": "Sheet", "cell": "B2", "field": "brand.logo"}]

    class DesignWithImage:
        output_format = "xlsx"
        xlsx_template = TemplateWithImage()

    class IssuanceWithImage:
        id = uuid4()
        design_version = DesignWithImage()
        input_data = {
            "brand": {
                "logo": (
                    "data:image/png;base64,"
                    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
                )
            }
        }

    generated = generate_document_file(
        IssuanceWithImage(),
        db=None,
        storage_provider=_Storage(_workbook_bytes()),
    )

    workbook = load_workbook(BytesIO(generated.content))
    assert len(workbook["Sheet"]._images) == 1
