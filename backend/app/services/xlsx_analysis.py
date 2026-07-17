import json
import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

TOKEN_PATTERN = re.compile(r"{{\s*([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)*)\s*}}")
REPEAT_DEFINED_NAME = "_docman_repeats"
IMAGE_DEFINED_NAME = "_docman_images"


@dataclass
class XlsxTemplateAnalysis:
    detected_sheets: list[dict]
    detected_tokens: list[str]
    image_slots: list[dict]
    validation_warnings: list[dict]


def analyze_xlsx_template(workbook_bytes: bytes, schema_tokens: set[str]) -> XlsxTemplateAnalysis:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=False, data_only=False)
    detected_sheets: list[dict] = []
    detected_tokens: list[str] = []
    validation_warnings: list[dict] = []
    seen_tokens: set[str] = set()
    repeat_specs = _load_repeat_specs(workbook, validation_warnings)
    image_slots = _load_image_slots(workbook, validation_warnings)

    for worksheet in workbook.worksheets:
        detected_sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "print_area": str(worksheet.print_area) if worksheet.print_area else None,
                "merged_ranges": [str(cell_range) for cell_range in worksheet.merged_cells.ranges],
            }
        )
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for token in TOKEN_PATTERN.findall(cell.value):
                    if token not in seen_tokens:
                        seen_tokens.add(token)
                        detected_tokens.append(token)
                    schema_token = _schema_token_for_cell(token, worksheet.title, cell.row, repeat_specs)
                    if schema_token not in schema_tokens:
                        validation_warnings.append(
                            {
                                "type": "unknown_schema_token",
                                "sheet": worksheet.title,
                                "cell": cell.coordinate,
                                "message": f"Token '{{{{{token}}}}}' is not defined by the document type",
                                "suggestion": "Add the field to the document type or replace the token",
                            }
                        )

    return XlsxTemplateAnalysis(
        detected_sheets=detected_sheets,
        detected_tokens=detected_tokens,
        image_slots=image_slots,
        validation_warnings=validation_warnings,
    )


def _load_repeat_specs(workbook, validation_warnings: list[dict]) -> list[dict]:
    defined_name = workbook.defined_names.get(REPEAT_DEFINED_NAME)
    if defined_name is None:
        return []

    specs: list[dict] = []
    try:
        destinations = list(defined_name.destinations)
    except Exception:
        validation_warnings.append({"type": "invalid_repeat_metadata", "message": "Invalid repeat metadata."})
        return specs

    for sheet_name, coordinate in destinations:
        if sheet_name not in workbook.sheetnames:
            validation_warnings.append({"type": "invalid_repeat_metadata", "message": "Repeat metadata references an unknown sheet."})
            continue
        raw_value = workbook[sheet_name][coordinate].value
        if raw_value in (None, ""):
            continue
        try:
            parsed = json.loads(raw_value)
        except (TypeError, ValueError):
            validation_warnings.append({"type": "invalid_repeat_metadata", "message": "Repeat metadata is not valid JSON."})
            continue
        if not isinstance(parsed, list):
            validation_warnings.append({"type": "invalid_repeat_metadata", "message": "Repeat metadata must be a list."})
            continue
        for item in parsed:
            if not isinstance(item, dict) or not {"sheet", "row", "list"} <= item.keys():
                validation_warnings.append({"type": "invalid_repeat_metadata", "message": "Repeat metadata entries require sheet, row, and list."})
                continue
            if item["sheet"] not in workbook.sheetnames:
                validation_warnings.append({"type": "invalid_repeat_metadata", "message": "Repeat metadata references an unknown sheet."})
                continue
            try:
                row = int(item["row"])
            except (TypeError, ValueError):
                validation_warnings.append({"type": "invalid_repeat_metadata", "message": "Repeat metadata row must be numeric."})
                continue
            specs.append({"sheet": item["sheet"], "row": row, "list": item["list"]})
    return specs


def _load_image_slots(workbook, validation_warnings: list[dict]) -> list[dict]:
    defined_name = workbook.defined_names.get(IMAGE_DEFINED_NAME)
    if defined_name is None:
        return []

    slots: list[dict] = []
    for sheet_name, coordinate in defined_name.destinations:
        if sheet_name not in workbook.sheetnames:
            validation_warnings.append({"type": "invalid_image_metadata", "message": "Image metadata references an unknown sheet."})
            continue
        raw_value = workbook[sheet_name][coordinate].value
        if raw_value in (None, ""):
            continue
        try:
            parsed = json.loads(raw_value)
        except (TypeError, ValueError):
            validation_warnings.append({"type": "invalid_image_metadata", "message": "Image metadata is not valid JSON."})
            continue
        if not isinstance(parsed, list):
            validation_warnings.append({"type": "invalid_image_metadata", "message": "Image metadata must be a list."})
            continue
        for item in parsed:
            if not isinstance(item, dict) or not {"sheet", "cell", "field"} <= item.keys():
                validation_warnings.append({"type": "invalid_image_metadata", "message": "Image metadata entries require sheet, cell, and field."})
                continue
            if item["sheet"] not in workbook.sheetnames:
                validation_warnings.append({"type": "invalid_image_metadata", "message": "Image metadata references an unknown sheet."})
                continue
            slots.append({"sheet": item["sheet"], "cell": item["cell"], "field": item["field"]})
    return slots


def _schema_token_for_cell(token: str, sheet_name: str, row_index: int, repeat_specs: list[dict]) -> str:
    if not token.startswith("item."):
        return token
    for spec in repeat_specs:
        if spec["sheet"] == sheet_name and spec["row"] == row_index:
            return f"{spec['list']}[].{token.removeprefix('item.')}"
    return token
