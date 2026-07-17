import copy
import json
from io import BytesIO
from typing import Any

from jinja2 import ChainableUndefined
from jinja2.sandbox import SandboxedEnvironment
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.services.xlsx_images import normalize_image_value


_JINJA_ENV = SandboxedEnvironment(autoescape=False, undefined=ChainableUndefined)
_REPEAT_DEFINED_NAME = "_docman_repeats"


def render_xlsx_template(
    workbook_bytes: bytes,
    payload: dict,
    image_values: dict | None = None,
) -> bytes:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    repeat_rows = _render_repeat_rows(workbook, payload)

    for worksheet in workbook.worksheets:
        skipped_rows = repeat_rows.get(worksheet.title, set())
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.row in skipped_rows:
                    continue
                _render_cell(cell, payload)

    _insert_images(workbook, image_values or {})

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def preview_xlsx_template(workbook_bytes: bytes, payload: dict) -> dict:
    rendered_bytes = render_xlsx_template(workbook_bytes, payload)
    workbook = load_workbook(BytesIO(rendered_bytes), data_only=False)
    sheets: list[dict] = []

    for worksheet in workbook.worksheets:
        cells = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells.append({"address": cell.coordinate, "value": _preview_cell_value(cell.value), "style": {}})
        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "merged_ranges": [str(cell_range) for cell_range in worksheet.merged_cells.ranges],
                "cells": cells,
            }
        )

    return {"sheets": sheets, "warnings": []}


def _render_repeat_rows(workbook, payload: dict) -> dict[str, set[int]]:
    rendered_rows: dict[str, set[int]] = {}
    offset_by_sheet: dict[str, int] = {}

    for spec in _load_repeat_specs(workbook):
        sheet_name = spec["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError("invalid_repeat_metadata")
        worksheet = workbook[sheet_name]
        row_index = int(spec["row"]) + offset_by_sheet.get(sheet_name, 0)
        items = _resolve_path(payload, spec["list"])
        if items is None:
            items = []
        if not isinstance(items, list):
            raise ValueError("repeat_list_must_be_array")

        _reject_merged_repeat_row(worksheet, row_index)
        rendered = _render_repeat_row(worksheet, row_index, payload, items)
        rendered_rows.setdefault(sheet_name, set()).update(rendered)
        offset_by_sheet[sheet_name] = offset_by_sheet.get(sheet_name, 0) + len(items) - 1

    return rendered_rows


def _load_repeat_specs(workbook) -> list[dict]:
    defined_name = workbook.defined_names.get(_REPEAT_DEFINED_NAME)
    if defined_name is None:
        return []

    specs: list[dict] = []
    for sheet_name, coordinate in defined_name.destinations:
        if sheet_name not in workbook.sheetnames:
            raise ValueError("invalid_repeat_metadata")
        worksheet = workbook[sheet_name]
        raw_value = worksheet[coordinate].value
        if raw_value in (None, ""):
            continue
        parsed = json.loads(raw_value)
        if not isinstance(parsed, list):
            raise ValueError("invalid_repeat_metadata")
        for item in parsed:
            if not isinstance(item, dict) or not {"sheet", "row", "list"} <= item.keys():
                raise ValueError("invalid_repeat_metadata")
            specs.append(item)

    return sorted(specs, key=lambda spec: (spec["sheet"], int(spec["row"])))


def _render_repeat_row(worksheet: Worksheet, row_index: int, payload: dict, items: list[dict]) -> set[int]:
    if not items:
        worksheet.delete_rows(row_index, 1)
        return set()

    template_cells = [_snapshot_cell(cell) for cell in worksheet[row_index]]
    template_height = worksheet.row_dimensions[row_index].height
    if len(items) > 1:
        worksheet.insert_rows(row_index + 1, len(items) - 1)

    rendered_rows: set[int] = set()
    for item_index, item in enumerate(items):
        target_row = row_index + item_index
        rendered_rows.add(target_row)
        if template_height is not None:
            worksheet.row_dimensions[target_row].height = template_height
        for template_cell in template_cells:
            target_cell = worksheet.cell(row=target_row, column=template_cell["column"])
            _apply_cell_snapshot(template_cell, target_cell)
            context = {**payload, "item": item}
            _render_cell(target_cell, context)

    return rendered_rows


def _snapshot_cell(cell) -> dict:
    return {
        "column": cell.column,
        "value": cell.value,
        "has_style": cell.has_style,
        "style": copy.copy(cell._style),
        "number_format": cell.number_format,
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "border": copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
        "protection": copy.copy(cell.protection),
    }


def _apply_cell_snapshot(snapshot: dict, target) -> None:
    target.value = snapshot["value"]
    if snapshot["has_style"]:
        target._style = copy.copy(snapshot["style"])
    target.number_format = snapshot["number_format"]
    target.font = copy.copy(snapshot["font"])
    target.fill = copy.copy(snapshot["fill"])
    target.border = copy.copy(snapshot["border"])
    target.alignment = copy.copy(snapshot["alignment"])
    target.protection = copy.copy(snapshot["protection"])


def _render_cell(cell, context: dict) -> None:
    if isinstance(cell.value, str) and "{{" in cell.value:
        cell.value = _escape_formula_text(_JINJA_ENV.from_string(cell.value).render(context))


def _escape_formula_text(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _insert_images(workbook, image_values: dict) -> None:
    for anchor, raw_value in image_values.items():
        if "!" not in anchor:
            raise ValueError("invalid_image_anchor")
        sheet_name, cell_coordinate = anchor.split("!", 1)
        if sheet_name not in workbook.sheetnames:
            raise ValueError("invalid_image_anchor")
        normalized = normalize_image_value(raw_value)
        image = OpenpyxlImage(BytesIO(normalized.content))
        image.anchor = cell_coordinate
        workbook[sheet_name].add_image(image)


def _preview_cell_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    return str(value)


def _reject_merged_repeat_row(worksheet: Worksheet, row_index: int) -> None:
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        del min_col, max_col
        if min_row <= row_index <= max_row:
            raise ValueError("unsupported_merge_in_repeat_range")


def _resolve_path(payload: dict, path: str) -> Any:
    current: Any = payload
    for segment in path.split("."):
        if not isinstance(current, dict) or segment not in current:
            return None
        current = current[segment]
    return current
